from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from datetime import datetime, timedelta
from env import *
from enum import Enum
import openpyxl as opx
import time
import requests
import os
import json
import platform

CARD_URL = f'https://api.trello.com/1/cards/{CARD_ID}/checklists'

API_TIMEOUT = 5
SAVE_TO_LOCAL = True
EYE_REST_NOTIFICATION = True

class Tags(Enum):
    undefined = (0, 'undefined')
    bci = (1, 'BCI')
    maliang = (2, 'Maliang')
    eat = (3, 'Eat')
    waste_time = (4, 'Waste time')
    read = (5, 'Read')
    entertainment = (6, 'Entertainment')
    logistics = (7, 'Logistics')
    sleep = (8, 'Sleep')
    
    def __str__(self):
        return self.value[1]
    
    def __int__(self):
        return self.value[0]

class TimerSettings:
    def __init__(self, settingsFilePath, 
                  lastRecordedDatetime = None, 
                  latestRecordFilePath = '', 
                  lastRecordWorksheetIndex = 0):
        self.lastRecordedDatetime = lastRecordedDatetime
        self.latestRecordFilePath = latestRecordFilePath
        self.lastRecordWorksheetIndex = lastRecordWorksheetIndex
        self.settingsFilePath = settingsFilePath
    
    def UpdateSettings(self, 
                       lastRecordedDatetime = None, 
                       latestRecordFilePath : str = '', 
                       lastRecordWorksheetIndex : int = -1):
        lrdt = lastRecordedDatetime if lastRecordedDatetime else self.lastRecordedDatetime
        settings = {
            'lastRecordedDatetime': GetStoreDatetimeString(lrdt),
            'latestRecordFilePath': latestRecordFilePath if latestRecordFilePath != '' else self.latestRecordFilePath,
            'lastRecordWorksheetIndex' : lastRecordWorksheetIndex if lastRecordWorksheetIndex != -1 else self.lastRecordWorksheetIndex, 
        }
        with open(self.settingsFilePath, 'w') as s:
            json.dump(settings, s, indent=4)
        
        self.lastRecordedDatetime = lrdt
        self.latestRecordFilePath = settings['latestRecordFilePath']
        self.lastRecordWorksheetIndex = settings['lastRecordWorksheetIndex']
    
    def FetchSettingsFromJson(self):
        with open(self.settingsFilePath) as s:
            # fetch settings data
            settings = json.load(s)
            self.lastRecordedDatetime = RetrieveDatetimeFromString(settings['lastRecordedDatetime'])
            self.latestRecordFilePath = settings['latestRecordFilePath']
            self.lastRecordWorksheetIndex = settings['lastRecordWorksheetIndex']
    
    def __str__(self):
        return str(vars(self))

class Timer:
    def __init__(self, debug = False):
        self.debug = debug
        #Saves
        self.dirPath = os.path.expanduser('~/Documents/WorkTimer')
        self.settingsFilePath = os.path.join(self.dirPath, 'settings.json')
        self.settings = TimerSettings(self.settingsFilePath)
        
        # Timer
        self.running = True
        self.lastRecordedDateString = None
        self.startTime = time.time()
        self.startDatetime = datetime.today()
        self.endDatetime = None
        
        # UI
        self.root = Tk()
        self.root.title("Work Timer")
        self.title = Entry(self.root)
        self.title.grid(row=0, column=0)
        self.title.focus()
        self.timeText = StringVar()
        self.timer = Label(self.root, textvariable=self.timeText, padx=100, pady=30).grid(row=1, column=0)
        self.root.bind('<Return>', lambda event: self.StopTimer())
        self.debugButton = None
        if self.debug:
            self.debugButton = Button(self.root, text="Debug", 
                           padx=100, pady=10,
                           command=self.Test).grid(row=3)
        self.checkButtonContainer = Frame(self.root, height=5, width=5)
        self.checkButtonContainer.grid(row = 4, column=0)
        self.tagsCheckbuttons = []
        self.tagsChecked = []
        r = 0
        for t in Tags:
            if t == Tags.undefined: # TODO: bad organization, fix
                self.tagsChecked.append(None)
                self.tagsCheckbuttons.append(None)
                continue
            self.tagsChecked.append(IntVar())
            self.tagsCheckbuttons.append(
                Checkbutton(self.checkButtonContainer, variable=self.tagsChecked[-1], 
                            onvalue=1, offvalue=0, text=str(t))
            )
            self.tagsCheckbuttons[-1].grid(row=r, column=0, sticky = W)
            r += 1 
        
        # launch
        self.StartUpCheckSaves()
        print(self.settings)
        self.UpdateTimer()
        self.root.mainloop()
    
    def StopTimer(self):
        if len(self.title.get()) == 0:
            return
        self.running = False
        self.endDatetime = datetime.today()
        
        if not self.debug:
            if SAVE_TO_LOCAL:
                self.RecordToLocal()
            else:
                self.RecordToTrello()
        
        self.title.delete(0, 'end')
        self.running = True
        self.startTime = time.time()
        self.startDatetime = datetime.today()
        for bt in self.tagsCheckbuttons[1:]:
            bt.deselect()
        self.UpdateTimer()
    
    def UpdateTimer(self):
        if not self.running: 
            return
        now = time.time()
        hours = TwoDigitNumber(str(int(now - self.startTime) // 3600))
        minutes = TwoDigitNumber(str((int(now - self.startTime) // 60) % 60))
        seconds = TwoDigitNumber(str(int(now - self.startTime) % 60))
        self.timeText.set(hours + ":" + minutes + ":" + seconds)
        self.root.after(1000, self.UpdateTimer)

# ----------------------- Trello option -------------------------------
    def GetMostRecentChecklistId(self):
        query = {
            'key': KEY,
            'token': TOKEN
        }
        response = requests.get(
            CARD_URL,
            params=query
        )
        return min(response.json(), key=lambda x: x['pos'])['id']
    
    def RecordToTrello(self):
        try:
            checklistId = self.GetMostRecentChecklistId()
            checklistUrl = f'https://api.trello.com/1/checklists/{checklistId}/checkItems'
            query = {
                'name': f'{TrelloFormatTime(self.timeText.get())} {self.title.get()}',
                'pos': 'bottom',
                'key': KEY,
                'token': TOKEN
            }
            response = requests.request(
                "POST",
                checklistUrl,
                params=query,
                timeout=API_TIMEOUT
            )
            messagebox.showinfo("You Are Done!", f"You worked for {self.timeText.get()}!")
            print(response.text)
            
            newRecordDateString = datetime.today().strftime('%Y-%m-%d')
            if(newRecordDateString != self.lastRecordedDateString):
                self.lastRecordedDateString = newRecordDateString
                query = {
                    'name': self.lastRecordedDateString,
                    'pos': 'bottom',
                    'key': KEY,
                    'token': TOKEN
                }
                response = requests.request(
                    "POST",
                    checklistUrl,
                    params=query,
                    timeout=API_TIMEOUT
                )
        except (requests.ConnectionError, requests.Timeout) as exception:
            messagebox.showinfo("No Connection", f"You worked for {self.timeText.get()}! Note this in your notepad. ")
    
# ----------------------- Local option -------------------------------
    def RecordToLocal(self):
        # check if need to create anything
        lrdt = self.settings.lastRecordedDatetime.date()
        if lrdt:
            if lrdt < GetClosestMonday():
                self.CreateNewRecordFile()
            elif lrdt and lrdt < datetime.today().date():
                self.CreateNewRecordSheet()
                
        self.AddRecord()
        messagebox.showinfo("You Are Done!", f"You worked for {self.timeText.get()}!")
    
    def StartUpCheckSaves(self):
        if not os.path.isdir(self.dirPath):
            os.mkdir(self.dirPath)
        if not os.path.isfile(self.settingsFilePath):
            self.CreateNewRecordFile()
        else:   # if the setting file already exist
            self.settings.FetchSettingsFromJson()
    
    def CreateNewRecordFile(self, date = None):
        date = date if date else datetime.today()
        todayString = date.strftime('%Y-%m-%d')
        recordFileName = f"{todayString}_Work_Record.xlsx"
        recordFilePath = os.path.join(self.dirPath, recordFileName)
        # https://docs.python.org/3/library/datetime.html#strftime-strptime-behavior
        newWorksheetName = f"{date.strftime('%m-%d %A')}"   # 06-09 Friday
        
        newRecordFile = opx.Workbook()
        worksheet = newRecordFile.active
        worksheet.title = newWorksheetName
        CreateHeader(worksheet)
        newRecordFile.save(filename=recordFilePath)
        
        self.settings.UpdateSettings(latestRecordFilePath=recordFilePath, 
                                lastRecordWorksheetIndex=0,
                                lastRecordedDatetime=datetime.today())
        
    def CreateNewRecordSheet(self, date = None):
        date = date if date else datetime.today()
        newWorksheetName = f"{date.strftime('%m-%d %A')}"   # 06-09 Friday
        record = opx.load_workbook(self.settings.latestRecordFilePath)
        newSheet = record.create_sheet(newWorksheetName)
        CreateHeader(newSheet)
        record.save(filename=self.settings.latestRecordFilePath)
        self.settings.UpdateSettings(lastRecordWorksheetIndex=(len(record.sheetnames)-1))
        
    def AddRecord(self):
        recordFile = opx.load_workbook(self.settings.latestRecordFilePath)
        recordFile.active = self.settings.lastRecordWorksheetIndex
        currentSheet = recordFile.active
        tagsChecked = [t for t in Tags if t != Tags.undefined and self.tagsChecked[int(t)].get() == 1]
        if len(tagsChecked) == 0:
            tagsChecked.append(Tags.undefined)

        record = [GetStoreDatetimeString(self.startDatetime),   # start time
                  GetStoreDatetimeString(self.endDatetime),     # end time
                  TrelloFormatTime(self.timeText.get()),        # duration
                  self.title.get(),                             # description
                  ','.join([str(t) for t in tagsChecked]),      # tags
                  ','.join([str(int(t)) for t in tagsChecked])] # tag-int
        currentSheet.append(record)
        recordFile.save(filename=self.settings.latestRecordFilePath)
        
        self.settings.UpdateSettings(lastRecordedDatetime=datetime.today())

    def Test(self):
        SendNotification('message of not', 'title')


def TwoDigitNumber(x):
    if len(x) < 2:
        return "0" + x
    else:
        return x
    
def TrelloFormatTime(x : str):
    (hours, minutes, seconds) = [int(i) for i in x.split(':')]
    minutes += int(seconds >= 30)
    hours = str(hours + minutes // 60)
    minutes = str(minutes % 60)
    return f'{TwoDigitNumber(hours)}:{TwoDigitNumber(minutes)}'

def GetClosestMonday():
    startDay = datetime.today()
    return (startDay - timedelta(days=startDay.weekday())).date()
    
def CreateHeader(worksheet):
    header = ['Start time', 'End time', 'Duration', 'Description', 'Tags', 'Tag-int']
    worksheet.append(header)
    
def GetStoreDatetimeString(t):
    return datetime.strftime(t, '%Y-%m-%d %H:%M:%S.%f')

def RetrieveDatetimeFromString(t):
    return datetime.strptime(t, '%Y-%m-%d %H:%M:%S.%f')

def SendNotification(message, title):
    # TODO: add notification for windows
    # https://www.pythongasm.com/desktop-notifications-with-python/
    plt = platform.system()
    
    if plt == 'Darwin':
        print("darwin")
        command = f'''
        osascript -e 'display notification "{message}" with title "{title}" sound name "Glass"'
        '''
    elif plt == 'Linux':
        command = f'''
        notify-send "{title}" "{message}"
        '''
    else:
        return
    print(command)
    os.system(command)

timer = Timer(debug = False)